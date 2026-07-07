# Created by: Roberto Serino
# Date: 2025-11-19 (modified 2025-12-11)
# Description: Analysis of working point results to compute S/N ratios
#              (modified to plot ALL good WP events: x = time, y = OptimumSN)

import os
import ROOT
import numpy as np
from openpyxl import Workbook
import argparse
import matplotlib.pyplot as plt
import plotly.graph_objects as go
import sys

ROOT.gErrorIgnoreLevel = ROOT.kWarning

ROOT.gROOT.SetBatch(True) # Don't print the canvas on screen

parser = argparse.ArgumentParser(description="Compute S/N ratios for working point analysis")

parser.add_argument("--inputdir", required=True,
                    help="Directory containing processed ROOT files")

parser.add_argument("--outputdir", required=True,
                    help="Directory where output Excel file will be saved")

parser.add_argument("--outputdirap", required=True,
                    help="Directory where png of the APs will be saved")


parser.add_argument("--outputdiranps", required=True,
                    help="Directory where png of the ANPSs will be saved")

parser.add_argument("--run", required=True,
                    help="Run number (string, e.g. 000045)")

parser.add_argument("--channels", nargs="+", type=int, required=False,
                    help="List of channels to process (optional)")

parser.add_argument("--workingpoints", nargs="+", type=float, required=False,
                    help="List of working points to process (optional)")

parser.add_argument("--wp_duration", type=float, default=60.00,
                    help="Duration of each working point in seconds (default: 60.00s)")

parser.add_argument("--offset", type=float, default=0.0,
                    help="Time offset to apply to working points in seconds (default: 0.0s)")

parser.add_argument("--pulsertype", required=False,
                    help="Type of pulser (e.g., LED, heater)")

args = parser.parse_args()

# Assign them to variables
ROOT_DIR = args.inputdir
OUTPUT_DIR = args.outputdir
OUTPUT_DIR_AP = args.outputdirap
OUTPUT_DIR_ANPS = args.outputdiranps
RUN_NUMBER = args.run
CHANNEL_LIST = args.channels   # None or a list of ints
WP_VALUES = args.workingpoints
WP_DURATION = args.wp_duration
OFFSET = args.offset
module_branch_name = args.pulsertype

# Guard: require WP_VALUES
if WP_VALUES is None or len(WP_VALUES) == 0:
    print("ERROR: No working points provided (use --workingpoints). Exiting.")
    sys.exit(1)

# Ensure output directory exists
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR_AP, exist_ok=True)
os.makedirs(OUTPUT_DIR_ANPS, exist_ok=True)

# ==========================
# Prepare data storage
# ==========================
data_sn = {}  # channel -> list of S/N values per working point
data_rise = {}  # channel -> list of rise times per working point

# Keep legacy names for downstream code, but they will hold "good events" (all WPs)
data_flagged_time = {}  # per channel: time_list for all good WP events
data_flagged_sn = {}    # per channel: S/N ratio list for all good WP events

MIN_POINTS = 20  # minimum number of points required
offset1 = (OFFSET * 0.87) / 2

num_wp = len(WP_VALUES)*2

# Precompute working point windows once (so Excel rows are consistent)
tstart = []
tend = []

for wp_index in range(num_wp):
    # center for this working point
    center = OFFSET + WP_DURATION / 2.0 + wp_index * (WP_DURATION + 25.0)
    start = center - offset1
    end = center + offset1
    tstart.append(start)
    tend.append(end)

print(f"\n=== Starting analysis for run {RUN_NUMBER} ===")
print(f"Working points: {WP_VALUES}")
print(f"Number of WPs: {num_wp}")
print(f"Pulser type: {module_branch_name}")
print(f"Input directory: {ROOT_DIR}")
print(f"Output directory: {OUTPUT_DIR}")
print(f"Output directory APs: {OUTPUT_DIR_AP}")
print(f"Output directory ANPSs: {OUTPUT_DIR_ANPS}")

# ==========================
# Process ROOT files
# ==========================
files_processed = 0
for filename in sorted(os.listdir(ROOT_DIR)):
    if not (filename.startswith("Processed") and filename.endswith(".root")):
        continue

    base, ext = os.path.splitext(filename)
    parts = base.split("_")
    if len(parts) < 4:
        print(f"Skipping unexpected filename format: {filename}")
        continue

    file_run = parts[-2]
    try:
        channel = int(parts[-1])
    except ValueError:
        print(f"Skipping file with non-integer channel part: {filename}")
        continue

    if channel == 13 or channel == 14 or channel == 41 or channel == 76 or channel == 94 or channel == 95:
        print(f"Skipping channel {channel} (known bad channel)")
    else:
        continue

    if file_run != RUN_NUMBER:
        continue

    # If user passed a limited channel list, skip others
    if CHANNEL_LIST is not None and channel not in CHANNEL_LIST:
        continue

    

    filepath = os.path.join(ROOT_DIR, filename)
    print(f"\n{'='*60}")
    print(f"Opening ROOT file: {filepath}")
    print(f"Channel: {channel}, Run: {file_run}")
    print(f"{'='*60}")

    root_file = ROOT.TFile.Open(filepath, "READ")
    if not root_file or root_file.IsZombie():
        print(f"  ERROR: Cannot open file {filepath}")
        continue

    # per-channel/per-file arrays
    wp_sn = [0.0] * num_wp
    wp_rise = [0.0] * num_wp

    # Collect ALL good events (from all WPs) for this channel
    # NOTE: these lists will contain one element per *good event* (across all WPs)
    channel_good_time = []
    channel_good_sn = []
    channel_good_rise = []
    
    files_processed += 1

    # Process each working point
    for wp_index in range(num_wp):

        if wp_index % 2 == 0:
            continue

        print(f"\n  Processing WP {wp_index}")


        ## Save all the APs in a folder:
        hist_name = f"averagepulse_ap_wp{wp_index}_trimmedAP"
        graph_name_anps = f"averagepowerspectrum_noise_wp{wp_index}_normalized"

        hist = root_file.Get(hist_name)
        if not hist:
            print(f"Histogram {hist_name} not found in {filename}")
            continue

        graph_anps = root_file.Get(graph_name_anps)
        if not graph_anps:
            print(f"Graph {graph_name_anps} not found in {filename}")
            continue
        
        wp_val = WP_VALUES[wp_index//2]

        if wp_index % 2 == 0:
            name = f"pos{wp_val}"
        else:
            name = f"neg{wp_val}"

        # Construct output PNG name
        png_name = f"AP_ch{channel}_WP{name}.jpg"
        png_path = os.path.join(OUTPUT_DIR_AP, png_name)

        png_anps_name = f"ANPS_ch{channel}_WP{name}.jpg"
        png_path_anps = os.path.join(OUTPUT_DIR_ANPS, png_anps_name)
        


        # Draw and save histogram
        c = ROOT.TCanvas()
        hist.Draw()
        c.SaveAs(png_path)

        c1 = ROOT.TCanvas()
        graph_anps.Draw()
        c1.SaveAs(png_path_anps)

        c.Close()
        
        # Try different tree name patterns
        tree_names_to_try = [
            f"optimumfilter__wp{wp_index}",
            f"optimumfilter_wp{wp_index}",
            f"optimumfilter{wp_index}",
            f"optimumfilter__{wp_index}"
        ]
        
        t_opt = None
        actual_tree_name = ""
        for tree_name in tree_names_to_try:
            t_opt = root_file.Get(tree_name)
            if t_opt:
                actual_tree_name = tree_name
                print(f"    Found optimum filter tree: {tree_name} with {t_opt.GetEntries()} entries")
                break
        
        if not t_opt:
            print(f"    ERROR: No optimum filter tree found for WP {wp_index}")
            continue

        t_time = root_file.Get("timestamp")
        t_rise = root_file.Get("risetime")
        
        # Get the baseline tree for RMS
        # t_base = root_file.Get("baseline")
        # if not t_base:
        #     print(f"    WARNING: baseline tree not found")
        
        # Check what branches are available
        # 
        
        branches = t_opt.GetListOfBranches()
        # for br in branches:
        #     print(f"      - {br.GetName()}")
        
        # Process entries
        n_entries = t_opt.GetEntries()
        print(f"    Processing {n_entries} entries...")
        
        # Set branch addresses
        time_val = np.zeros(1, dtype=float)
        sn_val = np.zeros(1, dtype=float)
        good_val = np.zeros(1, dtype=bool)
        rise_val = np.zeros(1, dtype=float)
        rms_val = np.zeros(1, dtype=float)
        
        # Try to set branch addresses
        try:
            t_time.SetBranchAddress("timefromstartrun", time_val)
        except Exception:
            print(f"    WARNING: Could not set branch address for timefromstartrun")
            # If timestamp tree exists but branch doesn't, skip WP (can't plot time)
            # Continue so we don't break whole script
            continue
        
        try:
            t_opt.SetBranchAddress("OptimumSN", sn_val)
        except Exception:
            print(f"    WARNING: Could not set branch address for OptimumSN")
            continue
        
        try:
            t_opt.SetBranchAddress("good", good_val)
        except Exception:
            print(f"    WARNING: Could not set branch address for good")
            good_val[0] = True  # Assume all are good if branch doesn't exist
        
        try:
            if t_rise:
                t_rise.SetBranchAddress("risetime", rise_val)
            else:
                rise_val[0] = 0.0
        except Exception:
            print(f"    WARNING: Could not set branch address for risetime")
            rise_val[0] = 0.0
        
        # Get RMS from baseline tree if available
        # baseline_rms = 1.0  # default
        # if t_base:
        #     try:
        #         t_base.SetBranchAddress("RMS", rms_val)
        #         if t_base.GetEntries() > 0:
        #             t_base.GetEntry(0)
        #             baseline_rms = rms_val[0]
        #             print(f"    Baseline RMS: {baseline_rms}")
        #     except Exception:
        #         print(f"    WARNING: Could not get RMS from baseline tree")
        
        # Collect good events for this WP
        good_sn_values = []
        good_rise_values = []
        
        for i in range(n_entries):
            t_opt.GetEntry(i)
            # read time (timestamp) from timestamp tree
            t_time.GetEntry(i)
            # read risetime from risetime tree if exists
            if t_rise:
                t_rise.GetEntry(i)
            
            # Skip if not good
            if not good_val[0]:
                continue
            
            # Skip invalid values
            if sn_val[0] <= 0 or sn_val[0] > 10000 or not np.isfinite(sn_val[0]):
                continue
            
            # Store for this WP (for statistics)
            good_sn_values.append(sn_val[0])
            good_rise_values.append(rise_val[0])
            
            # Store for overall scatter plot (ALL good events from ALL WPs)
            channel_good_time.append(time_val[0])
            channel_good_sn.append(sn_val[0])
            channel_good_rise.append(rise_val[0])
        
        print(f"    Found {len(good_sn_values)} good events in WP {wp_index}")
        
        # Calculate statistics for this WP
        if len(good_sn_values) >= MIN_POINTS:
            sn_array = np.array(good_sn_values)
            rise_array = np.array(good_rise_values)
            
            # Clean outliers for S/N
            med_sn = np.median(sn_array)
            mad_sn = np.median(np.abs(sn_array - med_sn))
            if mad_sn > 0:
                mask_sn = np.abs(sn_array - med_sn) / mad_sn < 3.0
                cleaned_sn = sn_array[mask_sn]
            else:
                cleaned_sn = sn_array
            
            if len(cleaned_sn) > 0:
                wp_sn[wp_index] = np.mean(cleaned_sn)
                print(f"    WP {wp_index} S/N: mean={wp_sn[wp_index]:.2f}, median={med_sn:.2f}, N={len(cleaned_sn)}")
            
            # Clean outliers for rise time
            if len(good_rise_values) >= MIN_POINTS:
                med_rise = np.median(rise_array)
                mad_rise = np.median(np.abs(rise_array - med_rise))
                if mad_rise > 0:
                    mask_rise = np.abs(rise_array - med_rise) / mad_rise < 3.0
                    cleaned_rise = rise_array[mask_rise]
                else:
                    cleaned_rise = rise_array
                
                if len(cleaned_rise) > 0:
                    wp_rise[wp_index] = np.mean(cleaned_rise)
                    print(f"    WP {wp_index} Rise time: {wp_rise[wp_index]:.3f} ms")
        else:
            print(f"    WARNING: Only {len(good_sn_values)} good events found in WP {wp_index} (need at least {MIN_POINTS})")
    
    # Store results for this channel
    data_sn[channel] = wp_sn
    data_rise[channel] = wp_rise
    
    # Store good events (all WPs) for scatter plot
    if channel_good_time:
        # keep original dictionary names so the rest of your script works unchanged
        data_flagged_time[channel] = channel_good_time
        data_flagged_sn[channel] = channel_good_sn
        
        # print(f"\n  Channel {channel}: Collected {len(channel_good_time)} good events (all WPs) total")
        
        # # Create scatter plot for this channel (matplotlib)
        # scatter_plot_dir = os.path.join(OUTPUT_DIR, "plots_flagged")
        # os.makedirs(scatter_plot_dir, exist_ok=True)
        
        # plt.figure(figsize=(10, 6))
        # plt.scatter(channel_good_time, channel_good_sn, s=10, alpha=0.7)
        # plt.xlabel("Time from start [s]")
        # plt.ylabel("OptimumSN (S/N ratio)")
        # plt.title(f"Channel {channel} - All good WP events (Run {RUN_NUMBER})")
        # plt.grid(True, alpha=0.3)
        # plt.tight_layout()
        
        # outfile_png = os.path.join(scatter_plot_dir, f"channel_{channel}_good_events.png")
        # plt.savefig(outfile_png, dpi=150)
        # plt.close()
        # print(f"    Saved scatter plot: {outfile_png}")
    else:
        print(f"\n  WARNING: No good WP events found for channel {channel}")
    
    root_file.Close()

print(f"\n{'='*60}")
print(f"Processed {files_processed} files")
print(f"Channels found: {sorted(data_sn.keys())}")
print(f"{'='*60}")

# ==========================
# Write Excel file
# ==========================
if not data_sn:
    print("ERROR: No data collected! Check your input files and parameters.")
    sys.exit(1)

wb = Workbook()
ws = wb.active
ws.title = f"Run_{RUN_NUMBER}"

# Headers
row = 1
col = 2
for time in tstart:
    ws.cell(row=row, column=col, value=time)
    col += 1

row = 2
col = 2
for time in tend:
    ws.cell(row=row, column=col, value=time)
    col += 1

row = 3
col = 2
for wp_label in WP_VALUES:
    ws.cell(row=row, column=col, value=wp_label)
    col += 1

# Data rows
row = 4
for channel in sorted(data_sn.keys()):
    # S/N values
    ws.cell(row=row, column=1, value=f"{channel} S/N")
    for col_idx, val in enumerate(data_sn[channel], start=2):
        ws.cell(row=row, column=col_idx, value=val)
    row += 1
    
    # Rise times (convert to milliseconds)
    ws.cell(row=row, column=1, value=f"{channel} RT")
    for col_idx, val in enumerate(data_rise[channel], start=2):
        ws.cell(row=row, column=col_idx, value=val * 1000)  # Convert to ms
    row += 1

outname = f"LoadCurves_run_{RUN_NUMBER}.xlsx"
outfile = os.path.join(OUTPUT_DIR, outname)
wb.save(outfile)

print(f"\nExcel file saved to:\n  {outfile}")

# ==========================
# Generate interactive HTML plot
# ==========================

html_dir = os.path.join(OUTPUT_DIR, "plots_flagged_html")
os.makedirs(html_dir, exist_ok=True)

fig = go.Figure()

# Add traces
sorted_channels = sorted(data_flagged_time.keys())
for channel in sorted_channels:
    # Skip empty data
    if len(data_flagged_time[channel]) == 0:
        continue

    fig.add_trace(go.Scatter(
        x=data_flagged_time[channel],
        y=data_flagged_sn[channel],
        mode='markers',
        name=f"Channel {channel}",
        visible=False,
        marker=dict(size=6, opacity=0.7)
    ))

# Make first channel visible
if len(fig.data) > 0:
    fig.data[0].visible = True
    default_channel = sorted_channels[0]
else:
    default_channel = ""

# Titles (without WP list)
full_title = f"<b>Run {RUN_NUMBER} – Good Flagged Events</b>"
new_title_base = f"<b>Run {RUN_NUMBER} – Channel {{channel}}</b>"

# Dropdown buttons (update only the title)
buttons = []
for i, channel in enumerate(sorted_channels):
    if i >= len(fig.data):
        continue  # skip channels that have no trace

    visible = [False] * len(fig.data)
    visible[i] = True
    new_title = new_title_base.format(channel=channel)

    buttons.append(dict(
        label=f"Channel {channel}",
        method="update",
        args=[
            {"visible": visible},
            {"title.text": new_title}
        ]
    ))

# Layout + annotation for WP list at bottom
fig.update_layout(
    title=dict(
        text=full_title,
        x=0.01,
        y=0.95,
        xanchor="left",
        yanchor="top",
        font=dict(size=20)
    ),
    annotations=[
        dict(
            text=f"Working points: {', '.join(map(str, WP_VALUES))}",
            x=0.5,
            y=-0.20,  # Below the x-axis
            xref="paper",
            yref="paper",
            showarrow=False,
            font=dict(size=14)
        )
    ],
    updatemenus=[dict(
        buttons=buttons,
        direction="down",
        showactive=True,
        x=0.98,
        xanchor="right",
        y=1.05,
        yanchor="top",
        bgcolor="white",
        bordercolor="lightgray",
        borderwidth=1,
        font=dict(size=14)
    )],
    xaxis_title="Time from start [s]",
    yaxis_title="S/N ratio",
    template="plotly_white",
    width=1000,
    height=600,
    margin=dict(t=120, b=120)  # add bottom margin so annotation fits
)

# Save HTML
outfile_html = os.path.join(html_dir, f"WPMeas_run_{RUN_NUMBER}.html")
fig.write_html(outfile_html)
print(f"\nInteractive HTML plot saved to:\n  {outfile_html}")
